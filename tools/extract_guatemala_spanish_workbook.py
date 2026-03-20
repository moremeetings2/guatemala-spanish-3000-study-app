#!/usr/bin/env python3

from __future__ import annotations

import json
import re
import sys
from pathlib import Path

from openpyxl import load_workbook


def slugify(value: str) -> str:
    slug = re.sub(r"[^a-z0-9]+", "-", value.lower()).strip("-")
    return slug or "item"


def clean(value):
    if value is None:
        return ""
    if isinstance(value, str):
        return value.strip()
    return value


def rows_after_header(sheet, header_label):
    header_seen = False
    for row in sheet.iter_rows(values_only=True):
        values = [clean(cell) for cell in row]
        if not any(values):
            continue
        if not header_seen:
            if values[0] == header_label:
                header_seen = True
            continue
        yield values


def build_main_words(sheet):
    entries = []
    for values in rows_after_header(sheet, "Rank"):
        if not values[0]:
            continue
        rank = int(values[0])
        band = clean(values[1])
        spanish = clean(values[2])
        common_forms = clean(values[3])
        english = clean(values[4])
        part_of_speech = clean(values[5])
        note = clean(values[7])
        lemma = clean(values[8])
        frequency_count = int(values[9]) if values[9] not in ("", None) else None
        entries.append(
            {
                "id": f"main-{rank:04d}",
                "type": "word",
                "sheet": "Spanish_3000",
                "rank": rank,
                "band": band,
                "spanish": spanish,
                "commonForms": common_forms,
                "english": english,
                "partOfSpeech": part_of_speech,
                "note": note,
                "lemma": lemma,
                "frequencyCount": frequency_count,
                "tags": [band, part_of_speech],
            }
        )
    return entries


def build_phrases(sheet):
    entries = []
    for index, values in enumerate(rows_after_header(sheet, "Spanish"), start=1):
        if not values[0]:
            continue
        spanish = clean(values[0])
        english = clean(values[1])
        when_to_use = clean(values[2])
        source = clean(values[3])
        entries.append(
            {
                "id": f"phrase-{index:03d}-{slugify(spanish)[:24]}",
                "type": "phrase",
                "sheet": "Coffee_Shop_Phrases",
                "spanish": spanish,
                "english": english,
                "context": when_to_use,
                "source": source,
                "tags": ["phrase", when_to_use] if when_to_use else ["phrase"],
            }
        )
    return entries


def build_bonus(sheet):
    entries = []
    for index, values in enumerate(rows_after_header(sheet, "Term"), start=1):
        if not values[0]:
            continue
        term = clean(values[0])
        meaning = clean(values[1])
        example = clean(values[2])
        source = clean(values[3])
        entries.append(
            {
                "id": f"bonus-{index:03d}-{slugify(term)[:24]}",
                "type": "bonus",
                "sheet": "Guatemala_Bonus",
                "spanish": term,
                "english": meaning,
                "note": example,
                "source": source,
                "tags": ["guatemala", "bonus"],
            }
        )
    return entries


def build_dashboard_metadata(sheet):
    metadata = {
        "title": "Guatemala Spanish 3000 Study Pack",
        "description": "",
        "stats": {},
    }
    for row in sheet.iter_rows(values_only=True):
        values = [clean(cell) for cell in row]
        if not any(values):
            continue
        if values[0] == "Guatemala Spanish 3000 Study Pack":
            metadata["title"] = values[0]
        elif isinstance(values[0], str) and values[0].startswith("Mobile-friendly"):
            metadata["description"] = values[0]
        elif len(values) > 1 and isinstance(values[0], str) and values[1] not in ("", None):
            metadata["stats"][values[0]] = values[1]
    return metadata


def main():
    if len(sys.argv) != 3:
        raise SystemExit("Usage: extract_guatemala_spanish_workbook.py <input.xlsx> <output.json>")

    input_path = Path(sys.argv[1]).expanduser()
    output_path = Path(sys.argv[2]).expanduser()
    workbook = load_workbook(input_path, read_only=True, data_only=True)

    dashboard = build_dashboard_metadata(workbook["Dashboard"])
    main_words = build_main_words(workbook["Spanish_3000"])
    phrases = build_phrases(workbook["Coffee_Shop_Phrases"])
    bonus = build_bonus(workbook["Guatemala_Bonus"])

    payload = {
        "meta": {
            "title": dashboard["title"],
            "description": dashboard["description"],
            "sourceWorkbook": input_path.name,
            "counts": {
                "words": len(main_words),
                "phrases": len(phrases),
                "bonus": len(bonus),
                "total": len(main_words) + len(phrases) + len(bonus),
            },
            "dashboardStats": dashboard["stats"],
        },
        "collections": {
            "mainWords": main_words,
            "coffeePhrases": phrases,
            "guatemalaBonus": bonus,
        },
    }

    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(json.dumps(payload, ensure_ascii=True, indent=2), encoding="utf-8")


if __name__ == "__main__":
    main()
